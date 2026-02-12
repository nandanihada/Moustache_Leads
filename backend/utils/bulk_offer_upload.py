"""
Bulk Offer Upload Utilities
Handles parsing and processing of spreadsheet uploads for bulk offer creation
Supports Excel (.xlsx, .xls), CSV, and Google Sheets
"""

from datetime import datetime, timedelta
from typing import List, Dict, Tuple, Any, Optional
import csv
import io
import re
import requests
from openpyxl import load_workbook

# Import offer constants for validation
from models.offer import (
    VALID_VERTICALS, 
    calculate_incentive_type,
    validate_vertical,
    map_category_to_vertical
)

# Field mapping from spreadsheet columns to database fields
SPREADSHEET_TO_DB_MAPPING = {
    'offer_id': 'campaign_id',  # In spreadsheet this is their campaign ID
    'campaign_id': 'campaign_id',
    'title': 'name',
    'name': 'name',
    'url': 'target_url',
    'target_url': 'target_url',
    'country': 'countries',
    'countries': 'countries',
    'payout': 'payout',
    'preview_url': 'preview_url',
    'preview url': 'preview_url',
    'image_url': 'image_url',
    'image url': 'image_url',
    'description': 'description',
    'platform': 'network',
    'name of platform': 'network',
    'network': 'network',
    'expiry': 'expiration_date',
    'expiration_date': 'expiration_date',
    # Vertical (replaces category) - supports both old and new column names
    'vertical': 'vertical',
    'category': 'vertical',  # Backward compatibility - maps to vertical
    'category of offer': 'vertical',
    'traffic_sources': 'affiliate_terms',
    'traffic sources': 'affiliate_terms',
    'device': 'device_targeting',
    'device_targeting': 'device_targeting',
    # NEW: Geo-restriction fields
    'non_access_url': 'non_access_url',
    'fallback_url': 'non_access_url',
    'blocked_url': 'non_access_url',
    'non access url': 'non_access_url',
    # NEW: Allowed countries for geo-restriction
    'allowed_countries': 'allowed_countries',
    'allowed countries': 'allowed_countries',
    'geo_restrict': 'allowed_countries',
    # NEW: Revenue sharing fields
    'percent': 'revenue_share_percent',
    'revenue_share_percent': 'revenue_share_percent',
    'revenue_percent': 'revenue_share_percent',
    'share_percent': 'revenue_share_percent',
    'revenue share': 'revenue_share_percent',
    # NEW: Payout model field (optional)
    'payout_model': 'payout_model',
    'payout model': 'payout_model',
    'model': 'payout_model',
    # NEW: Approval workflow fields
    'approval_type': 'approval_type',
    'approval type': 'approval_type',
    'auto_approve_delay': 'auto_approve_delay',
    'auto approve delay': 'auto_approve_delay',
    'approval_delay': 'auto_approve_delay',
    'approval delay': 'auto_approve_delay',
    'require_approval': 'require_approval',
    'require approval': 'require_approval',
    'approval_required': 'require_approval',
    # NEW: Offerwall visibility
    'show_in_offerwall': 'show_in_offerwall',
    'show in offerwall': 'show_in_offerwall',
    'offerwall': 'show_in_offerwall',
    'visible': 'show_in_offerwall',
    'visible_in_offerwall': 'show_in_offerwall',
}

# Required fields that must be present in spreadsheet
# Note: Either 'payout' OR 'revenue_share_percent' must be present (validated separately)
REQUIRED_FIELDS = [
    'campaign_id',  # offer_id in spreadsheet
    'name',         # title in spreadsheet
    'target_url',   # url in spreadsheet
    'countries',    # country in spreadsheet
    'description',
    'network'       # platform in spreadsheet
]

# Default values for optional fields
DEFAULT_VALUES = {
    'preview_url': 'https://www.google.com',  # Default preview URL if not provided
    'image_url': '',  # Will be set dynamically with random image
    'vertical': 'Lifestyle',  # NEW: Default vertical (replaces category)
    'affiliate_terms': 'Social, content, and direct traffic allowed.\nNo spam, incent abuse, or automation.\nFollow platform and advertiser rules.\nInvalid traffic will be rejected.',
    'status': 'active',  # Bulk uploaded offers are active by default (lowercase!)
    'offer_type': 'CPA',
    'currency': 'USD',
    'device_targeting': 'all',  # Default: all devices allowed
    'connection_type': 'all',
    'timezone': 'UTC',
    'access_type': 'public',
    'affiliates': 'all',
    'creative_type': 'image',
    'allowed_traffic_types': ['email', 'search', 'display'],
    'disallowed_traffic_types': ['adult', 'fraud'],
    'languages': [],
    'tags': [],
    'keywords': [],
    # NEW: Default values for new fields
    'allowed_countries': [],  # Empty = no geo-restriction
    'non_access_url': '',  # Empty = use system default
    'revenue_share_percent': 0,  # 0 = fixed payout
    'incentive_type': 'Incent',  # Default for fixed payout
    # NEW: Approval workflow defaults
    'approval_type': 'auto_approve',  # auto_approve, time_based, manual
    'auto_approve_delay': 0,  # Minutes delay for time_based approval
    'require_approval': False,  # Whether approval is required
    # NEW: Offerwall visibility
    'show_in_offerwall': True,  # Whether to show in offerwall (default: yes)
}

# List of real offer images to use as defaults
DEFAULT_OFFER_IMAGES = [
    'https://images.unsplash.com/photo-1607083206869-4c7672e72a8a?w=300&h=200&fit=crop',  # Shopping/ecommerce
    'https://images.unsplash.com/photo-1556742049-0cfed4f6a45d?w=300&h=200&fit=crop',  # Finance/money
    'https://images.unsplash.com/photo-1553729459-efe14ef6055d?w=300&h=200&fit=crop',  # Gaming
    'https://images.unsplash.com/photo-1526304640581-d334cdbbf45e?w=300&h=200&fit=crop',  # Crypto/tech
    'https://images.unsplash.com/photo-1563013544-824ae1b704d3?w=300&h=200&fit=crop',  # Travel
    'https://images.unsplash.com/photo-1498049794561-7780e7231661?w=300&h=200&fit=crop',  # Survey/form
    'https://images.unsplash.com/photo-1488590528505-98d2b5aba04b?w=300&h=200&fit=crop',  # Technology
    'https://images.unsplash.com/photo-1559526324-4b87b5e36e44?w=300&h=200&fit=crop',  # Dating/social
    'https://images.unsplash.com/photo-1579621970563-ebec7560ff3e?w=300&h=200&fit=crop',  # Health/fitness
    'https://images.unsplash.com/photo-1454165804606-c3d57bc86b40?w=300&h=200&fit=crop',  # Business/insurance
]

# Currency symbols for payout parsing
CURRENCY_SYMBOLS = {
    '$': 'USD',
    '€': 'EUR',
    '₹': 'INR',
    '£': 'GBP',
    '¥': 'JPY',
    '₽': 'RUB',
    'R$': 'BRL',
    'C$': 'CAD',
    'A$': 'AUD',
    'CHF': 'CHF',
    'kr': 'SEK',
    'zł': 'PLN',
    '₪': 'ILS',
    '₩': 'KRW',
    '฿': 'THB',
    '₫': 'VND',
    'Rp': 'IDR',
    'RM': 'MYR',
    '₱': 'PHP',
    'S$': 'SGD',
    'HK$': 'HKD',
    'NT$': 'TWD',
    'R': 'ZAR',
    'د.إ': 'AED',
    'SR': 'SAR',
    'QR': 'QAR',
    'BD': 'BHD',
    'KD': 'KWD',
    'OMR': 'OMR',
    '₤': 'GBP',  # Alternative pound symbol
}

# Currency indicators for validation
CURRENCY_INDICATORS = ['$', '€', '£', '₹', '¥', '₽']


def parse_payout_value(payout_str: str) -> Tuple[float, float, str]:
    """
    Parse payout value, detecting if it's a percentage or fixed amount with currency.
    
    Args:
        payout_str: Payout value from spreadsheet (e.g., "50", "50%", "$42", "€30", "₹100", "£25")
        
    Returns:
        Tuple of (fixed_payout, revenue_share_percent, currency)
        - If percentage: (0, percent_value, 'USD')
        - If fixed with currency: (payout_value, 0, detected_currency)
        - If fixed without currency: (payout_value, 0, 'USD')
    """
    payout_str = str(payout_str).strip()
    
    # Check if it's a percentage
    if '%' in payout_str:
        # Percentage payout - extract number
        percent = float(payout_str.replace('%', '').strip())
        return 0, percent, 'USD'
    
    # Detect currency symbol
    detected_currency = 'USD'  # Default
    numeric_part = payout_str
    
    # Check for currency symbols at the beginning or end
    for symbol, currency_code in CURRENCY_SYMBOLS.items():
        if payout_str.startswith(symbol):
            detected_currency = currency_code
            numeric_part = payout_str[len(symbol):].strip()
            break
        elif payout_str.endswith(symbol):
            detected_currency = currency_code
            numeric_part = payout_str[:-len(symbol)].strip()
            break
    
    # Remove any remaining non-numeric characters except decimal point
    numeric_part = re.sub(r'[^\d.]', '', numeric_part)
    
    # Parse the numeric value
    try:
        payout_value = float(numeric_part) if numeric_part else 0
    except ValueError:
        payout_value = 0
    
    return payout_value, 0, detected_currency


def normalize_column_name(column: str) -> str:
    """Normalize column names to lowercase and remove extra spaces"""
    return column.strip().lower()



def parse_excel_file(file_path: str) -> Tuple[List[Dict[str, Any]], Optional[str]]:
    """
    Parse Excel file (.xlsx, .xls) and return list of row dictionaries
    
    Args:
        file_path: Path to Excel file
        
    Returns:
        Tuple of (list of row dicts, error message if any)
    """
    try:
        workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
        sheet = workbook.active
        
        # Get headers from first row
        headers = []
        first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        for cell in first_row:
            if cell:
                headers.append(normalize_column_name(str(cell)))
            else:
                headers.append(None)
        
        # Parse data rows
        rows = []
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            row_data = {}
            has_data = False
            
            for idx, value in enumerate(row):
                if idx < len(headers) and headers[idx]:
                    # Convert value to string and clean it
                    if value is not None:
                        # Handle different data types
                        if isinstance(value, (int, float)):
                            row_data[headers[idx]] = str(value).strip()
                        else:
                            row_data[headers[idx]] = str(value).strip()
                        
                        if row_data[headers[idx]]:  # Check if not empty after stripping
                            has_data = True
            
            # Only add row if it has at least one non-empty value
            if has_data:
                row_data['_row_number'] = row_idx
                rows.append(row_data)
        
        workbook.close()
        return rows, None
        
    except Exception as e:
        import logging
        logging.error(f"Excel parsing error: {str(e)}", exc_info=True)
        return [], f"Error parsing Excel file: {str(e)}"


def parse_csv_file(file_path: str) -> Tuple[List[Dict[str, Any]], Optional[str]]:
    """
    Parse CSV file and return list of row dictionaries
    
    Args:
        file_path: Path to CSV file
        
    Returns:
        Tuple of (list of row dicts, error message if any)
    """
    try:
        rows = []
        with open(file_path, 'r', encoding='utf-8-sig') as csvfile:
            # Try to detect delimiter
            sample = csvfile.read(1024)
            csvfile.seek(0)
            
            sniffer = csv.Sniffer()
            try:
                dialect = sniffer.sniff(sample)
            except:
                dialect = csv.excel
            
            reader = csv.DictReader(csvfile, dialect=dialect)
            
            # Normalize headers
            normalized_fieldnames = [normalize_column_name(field) for field in reader.fieldnames]
            
            for row_idx, row in enumerate(reader, start=2):  # Start at 2 since row 1 is headers
                normalized_row = {}
                for original_key, value in row.items():
                    normalized_key = normalize_column_name(original_key)
                    normalized_row[normalized_key] = value
                
                # Skip empty rows
                if any(normalized_row.values()):
                    normalized_row['_row_number'] = row_idx
                    rows.append(normalized_row)
        
        return rows, None
        
    except Exception as e:
        return [], f"Error parsing CSV file: {str(e)}"


def fetch_google_sheet(sheet_url: str) -> Tuple[List[Dict[str, Any]], Optional[str]]:
    """
    Fetch data from Google Sheets URL (public sheet export as CSV)
    
    Args:
        sheet_url: Google Sheets URL
        
    Returns:
        Tuple of (list of row dicts, error message if any)
    """
    try:
        # Convert Google Sheets URL to CSV export URL
        if 'docs.google.com/spreadsheets' not in sheet_url:
            return [], "Invalid Google Sheets URL"
        
        # Extract sheet ID
        sheet_id_match = re.search(r'/d/([a-zA-Z0-9-_]+)', sheet_url)
        if not sheet_id_match:
            return [], "Could not extract sheet ID from URL"
        
        sheet_id = sheet_id_match.group(1)
        
        # Extract gid (sheet tab ID) if present
        gid_match = re.search(r'[#&]gid=([0-9]+)', sheet_url)
        gid = gid_match.group(1) if gid_match else '0'
        
        # Build CSV export URL
        csv_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"
        
        # Fetch CSV data
        response = requests.get(csv_url, timeout=30)
        response.raise_for_status()
        
        # Parse CSV content
        csv_content = response.content.decode('utf-8-sig')
        csv_file = io.StringIO(csv_content)
        
        reader = csv.DictReader(csv_file)
        rows = []
        
        for row_idx, row in enumerate(reader, start=2):
            normalized_row = {}
            for original_key, value in row.items():
                normalized_key = normalize_column_name(original_key)
                normalized_row[normalized_key] = value
            
            # Skip empty rows
            if any(normalized_row.values()):
                normalized_row['_row_number'] = row_idx
                rows.append(normalized_row)
        
        return rows, None
        
    except requests.exceptions.RequestException as e:
        return [], f"Error fetching Google Sheet: {str(e)}. Make sure the sheet is publicly accessible."
    except Exception as e:
        return [], f"Error parsing Google Sheet: {str(e)}"


def map_spreadsheet_to_db(row_data: Dict[str, Any]) -> Dict[str, Any]:
    """
    Map spreadsheet column names to database field names
    
    Args:
        row_data: Dictionary with spreadsheet column names
        
    Returns:
        Dictionary with database field names
    """
    mapped_data = {}
    
    for spreadsheet_col, db_field in SPREADSHEET_TO_DB_MAPPING.items():
        if spreadsheet_col in row_data and row_data[spreadsheet_col]:
            value = row_data[spreadsheet_col]
            
            # Handle country field - convert to array
            if db_field == 'countries' and value:
                # Split by comma if multiple countries
                if isinstance(value, str):
                    mapped_data[db_field] = [c.strip().upper() for c in value.split(',') if c.strip()]
                else:
                    mapped_data[db_field] = [str(value).strip().upper()]
            # Handle revenue_share_percent - remove % sign if present, detect currency values
            elif db_field == 'revenue_share_percent' and value:
                value_str = str(value).strip()
                # Check if this looks like a currency value (has $, €, etc.)
                is_currency_value = any(symbol in value_str for symbol in CURRENCY_INDICATORS)
                
                if is_currency_value:
                    # This is actually a payout value, not a percentage
                    # Store it temporarily, will be handled in validation
                    mapped_data[db_field] = value_str
                else:
                    # Remove % sign and store
                    if '%' in value_str:
                        value_str = value_str.replace('%', '').strip()
                    mapped_data[db_field] = value_str
            else:
                mapped_data[db_field] = value
    
    return mapped_data



def apply_default_values(row_data: Dict[str, Any]) -> Dict[str, Any]:
    """
    Apply default values for optional fields and auto-calculate incentive_type
    
    Args:
        row_data: Offer data dictionary
        
    Returns:
        Dictionary with defaults applied
    """
    import random
    import logging
    
    logger = logging.getLogger(__name__)
    
    result = row_data.copy()
    
    # Debug: Log if payout_model is present
    if 'payout_model' in result:
        logger.info(f"payout_model found in row data: {result['payout_model']}")
    
    # Apply default expiration date (90 days / 3 months from now)
    if not result.get('expiration_date'):
        expiry_date = datetime.utcnow() + timedelta(days=90)
        result['expiration_date'] = expiry_date.strftime('%Y-%m-%d')
    
    # Apply random image if not provided
    if not result.get('image_url'):
        result['image_url'] = random.choice(DEFAULT_OFFER_IMAGES)
    
    # Handle payout - detect percentage, currency, and extract revenue_share_percent
    if 'payout' in result and result['payout']:
        fixed_payout, revenue_share, detected_currency = parse_payout_value(str(result['payout']))
        if revenue_share > 0:
            result['payout'] = fixed_payout  # Will be 0 for percentage-based
            result['revenue_share_percent'] = revenue_share
        else:
            result['payout'] = fixed_payout
            # Only set to 0 if not already set from percent column
            if 'revenue_share_percent' not in result:
                result['revenue_share_percent'] = 0
        
        # Set currency if detected and not already set
        if 'currency' not in result or not result['currency']:
            result['currency'] = detected_currency
    elif 'revenue_share_percent' in result and result['revenue_share_percent']:
        # If only percent is provided, ensure payout is 0
        if 'payout' not in result or not result['payout']:
            result['payout'] = 0
    else:
        # Neither payout nor percent provided - set defaults
        if 'payout' not in result:
            result['payout'] = 0
        if 'revenue_share_percent' not in result:
            result['revenue_share_percent'] = 0

    # Handle vertical field - validate and normalize, or auto-detect from description
    if 'vertical' in result and result['vertical'] and result['vertical'].lower() != 'lifestyle':
        is_valid, vertical_result = validate_vertical(result['vertical'])
        if is_valid:
            result['vertical'] = vertical_result
        else:
            # Try mapping from old category value
            result['vertical'] = map_category_to_vertical(result['vertical'])
    else:
        # Auto-detect vertical from name and description
        from models.offer import detect_vertical_from_text
        offer_name = result.get('name', '')
        offer_description = result.get('description', '')
        result['vertical'] = detect_vertical_from_text(offer_name, offer_description)
    
    # Auto-calculate incentive_type based on revenue_share_percent
    revenue_share_percent = result.get('revenue_share_percent', 0)
    result['incentive_type'] = calculate_incentive_type(revenue_share_percent)
    
    # Handle allowed_countries for geo-restriction
    if 'allowed_countries' in result:
        allowed_countries = result['allowed_countries']
        if isinstance(allowed_countries, str):
            result['allowed_countries'] = [c.strip().upper() for c in allowed_countries.split(',') if c.strip()]
    
    # Handle approval workflow settings
    approval_type = result.get('approval_type', 'auto_approve')
    auto_approve_delay = result.get('auto_approve_delay', 0)
    require_approval = result.get('require_approval', False)
    
    # Normalize approval_type values
    if isinstance(approval_type, str):
        approval_type = approval_type.lower().strip()
        if approval_type in ['direct', 'instant', 'immediate', 'auto', 'auto_approve']:
            approval_type = 'auto_approve'
        elif approval_type in ['time', 'time_based', 'timed', 'delay', 'delayed']:
            approval_type = 'time_based'
        elif approval_type in ['manual', 'admin', 'approval']:
            approval_type = 'manual'
        else:
            approval_type = 'auto_approve'  # Default
    
    # Parse auto_approve_delay (convert to minutes)
    if auto_approve_delay:
        try:
            delay_str = str(auto_approve_delay).lower().strip()
            # Handle formats like "60", "60m", "1h", "1 hour"
            if 'h' in delay_str or 'hour' in delay_str:
                hours = float(re.sub(r'[^\d.]', '', delay_str))
                auto_approve_delay = int(hours * 60)
            elif 'd' in delay_str or 'day' in delay_str:
                days = float(re.sub(r'[^\d.]', '', delay_str))
                auto_approve_delay = int(days * 24 * 60)
            else:
                auto_approve_delay = int(float(re.sub(r'[^\d.]', '', delay_str)))
        except (ValueError, TypeError):
            auto_approve_delay = 0
    
    # Parse require_approval (convert to boolean)
    if isinstance(require_approval, str):
        require_approval = require_approval.lower().strip() in ['true', 'yes', '1', 'on']
    else:
        require_approval = bool(require_approval)
    
    # Auto-set require_approval based on approval_type
    if approval_type in ['time_based', 'manual']:
        require_approval = True
    
    # Build approval_settings object
    result['approval_settings'] = {
        'type': approval_type,
        'require_approval': require_approval,
        'auto_approve_delay': auto_approve_delay,
        'approval_message': '',
        'max_inactive_days': 0
    }
    
    # Set affiliates to 'request' if approval is required
    if require_approval or approval_type in ['time_based', 'manual']:
        result['affiliates'] = 'request'
    
    # Store individual fields for backward compatibility
    result['approval_type'] = approval_type
    result['auto_approve_delay'] = auto_approve_delay
    result['require_approval'] = require_approval
    
    # Apply other defaults
    for field, default_value in DEFAULT_VALUES.items():
        if field == 'image_url':  # Skip image_url as we handled it above
            continue
        if field == 'incentive_type':  # Skip - already calculated above
            continue
        if field not in result or not result[field]:
            result[field] = default_value
    
    return result



def validate_spreadsheet_data(rows: List[Dict[str, Any]], store_missing: bool = True) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]]]:
    """
    Validate spreadsheet data and apply field mapping and defaults.
    Offers with missing required fields are stored in Missing Offers collection.
    
    Args:
        rows: List of raw spreadsheet row dictionaries
        store_missing: Whether to return missing offers separately (True) or as errors (False)
        
    Returns:
        Tuple of (valid rows list, error rows list, missing offers list)
        - valid_rows: Offers that passed all validation
        - error_rows: Offers with critical errors (invalid data format)
        - missing_offers: Offers with missing required fields (to be stored in Missing Offers)
    """
    valid_rows = []
    error_rows = []
    missing_offers_rows = []
    
    for row in rows:
        row_number = row.get('_row_number', 'Unknown')
        errors = []
        missing_fields = []
        
        # Map spreadsheet columns to database fields
        mapped_data = map_spreadsheet_to_db(row)
        
        # Debug logging
        import logging
        logger = logging.getLogger(__name__)
        logger.info(f"Row {row_number} mapped data: {mapped_data}")

        # Check required fields - track which are missing
        for required_field in REQUIRED_FIELDS:
            if required_field not in mapped_data or not mapped_data[required_field]:
                missing_fields.append(required_field)
        
        # Special validation: Either 'payout' OR 'revenue_share_percent' must be present
        has_payout = 'payout' in mapped_data and mapped_data['payout']
        has_percent = 'revenue_share_percent' in mapped_data and mapped_data['revenue_share_percent']
        
        if not has_payout and not has_percent:
            missing_fields.append('payout')
        
        # If only percent is provided, set payout to 0
        if has_percent and not has_payout:
            mapped_data['payout'] = 0
            # Remove payout from missing fields if it was added
            if 'payout' in missing_fields:
                missing_fields.remove('payout')

        # Validate payout is numeric (handle percentage format like "50%" and currency symbols like "$42")
        if 'payout' in mapped_data and mapped_data['payout']:
            payout_str = str(mapped_data['payout']).strip()
            try:
                # Try to parse with currency detection
                fixed_payout, revenue_share, currency = parse_payout_value(payout_str)
                # If parsing succeeds, the value is valid
            except (ValueError, TypeError):
                errors.append(f"Invalid payout value: {mapped_data['payout']} (must be numeric, percentage like '50%', or with currency symbol like '$42')")
        
        # Validate vertical if provided
        if 'vertical' in mapped_data and mapped_data['vertical']:
            is_valid, result = validate_vertical(mapped_data['vertical'])
            if not is_valid:
                # Check if it can be mapped from old category
                mapped_vertical = map_category_to_vertical(mapped_data['vertical'])
                if mapped_vertical == 'Lifestyle' and mapped_data['vertical'].lower() not in ['lifestyle', 'general', '']:
                    errors.append(f"Invalid vertical '{mapped_data['vertical']}'. Must be one of: {', '.join(VALID_VERTICALS)}")
        
        # Validate revenue_share_percent if provided
        if 'revenue_share_percent' in mapped_data and mapped_data['revenue_share_percent']:
            percent_str = str(mapped_data['revenue_share_percent']).strip()
            
            # Check if this looks like a payout value (has currency symbol like $, €, etc.)
            is_currency_value = any(symbol in percent_str for symbol in CURRENCY_INDICATORS)
            
            if is_currency_value:
                # This is actually a payout value, not a percentage - move it to payout
                if not mapped_data.get('payout'):
                    mapped_data['payout'] = percent_str
                    # Remove payout from missing fields if it was added
                    if 'payout' in missing_fields:
                        missing_fields.remove('payout')
                mapped_data['revenue_share_percent'] = 0
            else:
                try:
                    # Remove % sign if present
                    percent = float(percent_str.replace('%', '').strip())
                    if percent < 0 or percent > 100:
                        errors.append(f"Invalid revenue_share_percent: {mapped_data['revenue_share_percent']} (must be 0-100)")
                except (ValueError, TypeError):
                    errors.append(f"Invalid revenue_share_percent: {mapped_data['revenue_share_percent']} (must be numeric)")
        
        # Validate URL format (allow macros like {user_id})
        if 'target_url' in mapped_data:
            # First, temporarily replace macros with placeholder for validation
            temp_url = str(mapped_data['target_url'])
            import re as regex_module
            # Replace all {macro} patterns with a valid placeholder
            temp_url = regex_module.sub(r'\{[^}]+\}', 'MACRO_PLACEHOLDER', temp_url)
            
            url_pattern = re.compile(
                r'^https?://'
                r'(?:(?:[A-Z0-9](?:[A-Z0-9-]{0,61}[A-Z0-9])?\.)+[A-Z]{2,6}\.?|'
                r'localhost|'
                r'\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})'
                r'(?::\d+)?'
                r'(?:/?|[/?]\S+)$', re.IGNORECASE)
            
            if not url_pattern.match(temp_url):
                errors.append(f"Invalid URL format: {mapped_data['target_url']}")
        
        # Determine where this row goes:
        # 1. If critical errors (invalid format) -> error_rows
        # 2. If missing required fields but no format errors -> missing_offers_rows
        # 3. If all good -> valid_rows
        
        if errors:
            # Critical errors - invalid data format
            error_rows.append({
                'row': row_number,
                'errors': errors,
                'data': row
            })
        elif missing_fields and store_missing:
            # Missing required fields - store in Missing Offers
            missing_offers_rows.append({
                'row': row_number,
                'missing_fields': missing_fields,
                'data': mapped_data,
                'raw_data': row
            })
        elif missing_fields:
            # Missing fields but store_missing=False - treat as error
            error_rows.append({
                'row': row_number,
                'errors': [f"Missing required field: {f}" for f in missing_fields],
                'data': row
            })
        else:
            # Apply defaults and add to valid rows
            validated_data = apply_default_values(mapped_data)
            validated_data['_row_number'] = row_number
            valid_rows.append(validated_data)
    
    return valid_rows, error_rows, missing_offers_rows



def bulk_create_offers(validated_data: List[Dict[str, Any]], created_by: str) -> Tuple[List[str], List[Dict[str, Any]], List[Dict[str, Any]]]:
    """
    Create multiple offers in the database.
    
    Args:
        validated_data: List of validated offer data dictionaries
        created_by: Username/ID of admin creating the offers
        
    Returns:
        Tuple of (list of created offer IDs, list of errors, empty list for backward compatibility)
    """
    from models.offer import Offer
    import logging
    
    logger = logging.getLogger(__name__)
    offer_model = Offer()
    created_offer_ids = []
    creation_errors = []
    
    for offer_data in validated_data:
        row_number = offer_data.pop('_row_number', 'Unknown')
        
        try:
            logger.info(f"Creating offer for row {row_number}: {offer_data}")
            
            # Create offer using the existing create_offer method
            created_offer, error = offer_model.create_offer(offer_data, created_by)
            
            if error:
                logger.error(f"Offer creation failed for row {row_number}: {error}")
                creation_errors.append({
                    'row': row_number,
                    'error': error,
                    'data': offer_data
                })
            else:
                logger.info(f"Successfully created offer {created_offer['offer_id']} for row {row_number}")
                created_offer_ids.append(created_offer['offer_id'])
                
        except Exception as e:
            creation_errors.append({
                'row': row_number,
                'error': f"Unexpected error: {str(e)}",
                'data': offer_data
            })
    
    return created_offer_ids, creation_errors, []


def check_inventory_gaps(offers_data: List[Dict[str, Any]], network: str = None) -> Dict[str, Any]:
    """
    Check a list of offers against inventory and identify gaps.
    This is the main function for inventory gap detection.
    
    Match Key = Name + Country + Platform (iOS/Android/Web) + Payout Model
    
    Args:
        offers_data: List of offer dictionaries to check
        network: Partner network name
        
    Returns:
        Dictionary with:
        - in_inventory: list of offers that exist
        - not_in_inventory: list of offers that don't exist (stored as missing)
        - stats: summary statistics
    """
    import logging
    logger = logging.getLogger(__name__)
    
    try:
        from services.missing_offer_service import MissingOfferService
        
        result = MissingOfferService.process_bulk_upload_for_gaps(
            offers=offers_data,
            source='sheet_upload',
            network=network
        )
        
        return {
            'in_inventory': result['existing_offers'],
            'not_in_inventory': result['missing_offers'],
            'stats': result['stats']
        }
        
    except ImportError:
        logger.error("MissingOfferService not available")
        return {
            'in_inventory': [],
            'not_in_inventory': [],
            'stats': {'total': len(offers_data), 'error': 'Service not available'}
        }
    except Exception as e:
        logger.error(f"Error checking inventory gaps: {e}")
        return {
            'in_inventory': [],
            'not_in_inventory': [],
            'stats': {'total': len(offers_data), 'error': str(e)}
        }
