"""Export Service for Offer Access Requests."""
import io
import logging
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from database import db_instance

logger = logging.getLogger(__name__)

COLORS = {"header_bg": "1F2937", "header_font": "FFFFFF", "section_bg": "374151", "section_font": "FFFFFF", "summary_title": "6366F1"}

STATUS_FILLS = {
    "approved": PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"),
    "rejected": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
    "in_review": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
    "pending": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
    "review": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
}

ALL_COLUMNS = {
    "offer_name": {"label": "Offer Name", "width": 30},
    "status": {"label": "Status", "width": 12},
    "publisher_username": {"label": "Publisher", "width": 20},
    "offer_network": {"label": "Network", "width": 18},
    "offer_category": {"label": "Vertical", "width": 16},
    "offer_countries": {"label": "Countries", "width": 20},
    "offer_payout": {"label": "Payout", "width": 12},
    "requested_at": {"label": "Requested At", "width": 18},
    "approved_at": {"label": "Approved At", "width": 18},
    "rejected_at": {"label": "Rejected At", "width": 18},
    "approved_by_username": {"label": "Approved By", "width": 16},
    "rejected_by_username": {"label": "Rejected By", "width": 16},
    "rejection_reason": {"label": "Rejection Reason", "width": 30},
    "rejection_category": {"label": "Rejection Category", "width": 20},
    "publisher_email": {"label": "Publisher Email", "width": 28},
    "offer_status": {"label": "Offer Status", "width": 14},
    "has_placement_proof": {"label": "Has Proof", "width": 12},
    "request_count": {"label": "Request Count", "width": 14},
    "total_requests": {"label": "Total Requests", "width": 14},
    "approved_count": {"label": "Approved Count", "width": 14},
    "rejected_count": {"label": "Rejected Count", "width": 14},
    "unique_users": {"label": "Unique Users", "width": 14},
    "offer_health": {"label": "Offer Health", "width": 14},
    "last_mail_sent": {"label": "Last Mail Sent", "width": 18},
    "approval_rate": {"label": "Approval Rate %", "width": 14},
    "unique_requesters": {"label": "Unique Requesters", "width": 16},
    "offer_total_requests": {"label": "Offer Total Requests", "width": 18},
    "days_since_request": {"label": "Days Since Request", "width": 18},
    "publisher_account_age": {"label": "Account Age", "width": 16},
    "publisher_clicks": {"label": "Pub Clicks", "width": 12},
    "publisher_conversions": {"label": "Pub Conversions", "width": 16},
    "message": {"label": "Message", "width": 30},
}


class ExportService:
    def __init__(self):
        self.requests_col = db_instance.get_collection("affiliate_requests")
        self.offers_col = db_instance.get_collection("offers")
        self.users_col = db_instance.get_collection("users")
        self.proofs_col = db_instance.get_collection("placement_proofs")

    def generate_export(self, config):
        tabs = config.get("tabs", ["approved"])
        columns = config.get("columns", list(ALL_COLUMNS.keys())[:8])
        group_by = config.get("group_by", "none")
        separate_sheets = config.get("separate_sheets_per_group", False)
        date_range = config.get("date_range")
        include_summary = config.get("include_summary", True)
        freeze_headers = config.get("freeze_headers", True)
        color_code_rows = config.get("color_code_rows", True)
        auto_fit = config.get("auto_fit_columns", True)
        wb = Workbook()
        wb.remove(wb.active)
        summary_data = {}
        for tab in tabs:
            data = self._fetch_tab_data(tab, date_range) if tab != "most_requested" else self._fetch_most_requested(self._build_date_filter(date_range))
            summary_data[tab] = {"count": len(data), "data": data}
            if group_by != "none" and separate_sheets:
                groups = self._group_data(data, group_by)
                for gn, gr in groups.items():
                    ws = wb.create_sheet(title=self._safe_sheet_name(f"{tab} - {gn}"))
                    self._write_sheet(ws, gr, columns, freeze_headers, color_code_rows, auto_fit)
            elif group_by != "none":
                ws = wb.create_sheet(title=self._safe_sheet_name(tab.replace("_", " ").title()))
                self._write_grouped_sheet(ws, self._group_data(data, group_by), columns, group_by, freeze_headers, color_code_rows, auto_fit)
            else:
                ws = wb.create_sheet(title=self._safe_sheet_name(tab.replace("_", " ").title()))
                self._write_sheet(ws, data, columns, freeze_headers, color_code_rows, auto_fit)
        if include_summary:
            self._create_summary_sheet(wb, summary_data, tabs, columns, config)
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def _build_date_filter(self, date_range):
        df = {}
        if date_range:
            f = self._parse_date(date_range.get("from"))
            t = self._parse_date(date_range.get("to"))
            if f: df["$gte"] = f
            if t: df["$lte"] = t.replace(hour=23, minute=59, second=59)
        return df

    def _parse_date(self, ds):
        if not ds: return None
        if isinstance(ds, datetime): return ds
        for fmt in ["%Y-%m-%dT%H:%M:%S.%fZ", "%Y-%m-%dT%H:%M:%SZ", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d"]:
            try: return datetime.strptime(ds, fmt)
            except: pass
        return None

    def _fetch_tab_data(self, tab, date_range=None):
        from bson import ObjectId
        date_filter = self._build_date_filter(date_range)
        status_map = {"approved": "approved", "rejected": "rejected", "in_review": ["pending", "review"], "all_requests": None}
        mapped = status_map.get(tab)
        q = {}
        if mapped is not None:
            q["status"] = {"$in": mapped} if isinstance(mapped, list) else mapped
        if date_filter:
            q["requested_at"] = date_filter
        reqs = list(self.requests_col.find(q).sort("requested_at", -1).limit(10000))
        if not reqs: return []
        offer_ids = list({r.get("offer_id") for r in reqs if r.get("offer_id")})
        user_ids = list({str(r.get("user_id")) for r in reqs if r.get("user_id")})
        # Offer cache
        offer_cache = {}
        if offer_ids and self.offers_col is not None:
            for o in self.offers_col.find({"offer_id": {"$in": offer_ids}}):
                offer_cache[o["offer_id"]] = o
        # User cache
        user_cache = {}
        if user_ids and self.users_col is not None:
            oids = []
            for uid in user_ids:
                try: oids.append(ObjectId(uid))
                except: pass
            if oids:
                for u in self.users_col.find({"_id": {"$in": oids}}):
                    user_cache[str(u["_id"])] = u
        # Proofs
        proof_set = set()
        if self.proofs_col is not None and user_ids:
            for doc in self.proofs_col.find({"user_id": {"$in": user_ids}}, {"user_id": 1, "offer_id": 1}):
                proof_set.add(f"{doc.get('user_id')}:{doc.get('offer_id')}")
        # Offer stats aggregation
        offer_stats = {}
        if offer_ids:
            for doc in self.requests_col.aggregate([{"$match": {"offer_id": {"$in": offer_ids}}}, {"$group": {"_id": "$offer_id", "total": {"$sum": 1}, "approved": {"$sum": {"$cond": [{"$eq": ["$status", "approved"]}, 1, 0]}}, "rejected": {"$sum": {"$cond": [{"$eq": ["$status", "rejected"]}, 1, 0]}}, "users": {"$addToSet": "$user_id"}}}]):
                oid = doc["_id"]
                t = doc["total"]
                a = doc["approved"]
                offer_stats[oid] = {"total": t, "approved": a, "rejected": doc["rejected"], "unique": len(doc.get("users", [])), "rate": round(a/t*100, 1) if t > 0 else 0}
        # Publisher mail
        mail_cache = {}
        email_col = db_instance.get_collection("email_activity_logs")
        send_col = db_instance.get_collection("offer_send_history")
        if user_ids and email_col is not None:
            for doc in email_col.find({"$or": [{"user_id": {"$in": user_ids}}, {"recipient_user_ids": {"$in": user_ids}}]}).sort("created_at", -1):
                for uid_m in ([doc.get("user_id", "")] + (doc.get("recipient_user_ids") or [])):
                    if uid_m in user_ids and uid_m not in mail_cache:
                        mail_cache[uid_m] = doc.get("created_at")
        if user_ids and send_col is not None:
            for doc in send_col.find({"$or": [{"user_id": {"$in": user_ids}}, {"recipient_user_ids": {"$in": user_ids}}]}).sort("created_at", -1):
                for uid_m in ([doc.get("user_id", "")] + (doc.get("recipient_user_ids") or [])):
                    if uid_m in user_ids and uid_m not in mail_cache:
                        mail_cache[uid_m] = doc.get("created_at")
        # Publisher clicks/conversions
        clicks_cache, conv_cache = {}, {}
        clicks_col = db_instance.get_collection("clicks")
        conv_col = db_instance.get_collection("conversions")
        if user_ids and clicks_col is not None:
            for doc in clicks_col.aggregate([{"$match": {"user_id": {"$in": user_ids}}}, {"$group": {"_id": "$user_id", "c": {"$sum": 1}}}]):
                clicks_cache[doc["_id"]] = doc["c"]
        if user_ids and conv_col is not None:
            for doc in conv_col.aggregate([{"$match": {"user_id": {"$in": user_ids}}}, {"$group": {"_id": "$user_id", "c": {"$sum": 1}}}]):
                conv_cache[doc["_id"]] = doc["c"]
        now = datetime.utcnow()
        results = []
        for r in reqs:
            oid = r.get("offer_id")
            offer = offer_cache.get(oid, {})
            uid = str(r.get("user_id", ""))
            usr = user_cache.get(uid, {})
            countries = offer.get("countries", [])
            if isinstance(countries, list): countries = ", ".join(countries[:5])
            req_at = r.get("requested_at")
            days_since = f"{(now - req_at).days} days" if isinstance(req_at, datetime) else ""
            created = usr.get("created_at")
            acct_age = f"{(now - created).days} days" if isinstance(created, datetime) else ""
            os = offer_stats.get(oid, {"total": 0, "approved": 0, "rejected": 0, "unique": 0, "rate": 0})
            results.append({"offer_name": offer.get("name", r.get("offer_name", "Unknown")), "status": r.get("status", ""), "publisher_username": usr.get("username", r.get("username", "")), "publisher_email": usr.get("email", r.get("email", "")), "offer_network": offer.get("network", ""), "offer_category": offer.get("category", offer.get("vertical", "")), "offer_countries": countries, "offer_payout": offer.get("payout", 0), "requested_at": req_at, "approved_at": r.get("approved_at"), "rejected_at": r.get("rejected_at"), "approved_by_username": r.get("approved_by_username", ""), "rejected_by_username": r.get("rejected_by_username", ""), "rejection_reason": r.get("rejection_reason", ""), "rejection_category": r.get("rejection_category", ""), "offer_status": offer.get("status", ""), "has_placement_proof": "Yes" if f"{uid}:{oid}" in proof_set else "No", "request_count": 1, "total_requests": os["total"], "approved_count": os["approved"], "rejected_count": os["rejected"], "unique_users": os["unique"], "offer_health": "", "approval_rate": f"{os['rate']}%", "unique_requesters": os["unique"], "offer_total_requests": os["total"], "days_since_request": days_since, "last_mail_sent": mail_cache.get(uid), "publisher_account_age": acct_age, "publisher_clicks": clicks_cache.get(uid, 0), "publisher_conversions": conv_cache.get(uid, 0), "message": r.get("message", "")})
        return results

    def _fetch_most_requested(self, date_filter):
        match = {"requested_at": date_filter} if date_filter else {}
        pipeline = [{"$match": match}] if match else []
        pipeline += [{"$group": {"_id": "$offer_id", "total": {"$sum": 1}, "approved": {"$sum": {"$cond": [{"$eq": ["$status", "approved"]}, 1, 0]}}, "rejected": {"$sum": {"$cond": [{"$eq": ["$status", "rejected"]}, 1, 0]}}, "users": {"$addToSet": "$user_id"}, "last_at": {"$max": "$requested_at"}, "last_user": {"$last": "$username"}}}, {"$sort": {"total": -1}}, {"$limit": 500}]
        raw = list(self.requests_col.aggregate(pipeline))
        oids = [r["_id"] for r in raw if r.get("_id")]
        offer_map = {}
        if oids and self.offers_col is not None:
            for o in self.offers_col.find({"offer_id": {"$in": oids}}):
                offer_map[o["offer_id"]] = o
        # Fetch offer-level clicks and conversions
        offer_clicks = {}
        offer_convs = {}
        clicks_col = db_instance.get_collection("clicks")
        conv_col = db_instance.get_collection("conversions")
        if oids and clicks_col is not None:
            for doc in clicks_col.aggregate([{"$match": {"offer_id": {"$in": oids}}}, {"$group": {"_id": "$offer_id", "c": {"$sum": 1}}}]):
                offer_clicks[doc["_id"]] = doc["c"]
        if oids and conv_col is not None:
            for doc in conv_col.aggregate([{"$match": {"offer_id": {"$in": oids}}}, {"$group": {"_id": "$offer_id", "c": {"$sum": 1}}}]):
                offer_convs[doc["_id"]] = doc["c"]
        now = datetime.utcnow()
        results = []
        for r in raw:
            oid = r["_id"]
            offer = offer_map.get(oid, {})
            countries = offer.get("countries", [])
            if isinstance(countries, list): countries = ", ".join(countries[:5])
            t, a, rej = r.get("total", 0), r.get("approved", 0), r.get("rejected", 0)
            uniq = len(r.get("users", []))
            rate = round(a/t*100, 1) if t > 0 else 0
            last_at = r.get("last_at")
            days = f"{(now - last_at).days} days" if isinstance(last_at, datetime) else ""
            results.append({"offer_name": offer.get("name", oid or "Unknown"), "status": "most_requested", "publisher_username": r.get("last_user", ""), "publisher_email": "", "offer_network": offer.get("network", ""), "offer_category": offer.get("category", offer.get("vertical", "")), "offer_countries": countries, "offer_payout": offer.get("payout", 0), "requested_at": last_at, "approved_at": None, "rejected_at": None, "approved_by_username": "", "rejected_by_username": "", "rejection_reason": "", "rejection_category": "", "offer_status": offer.get("status", ""), "has_placement_proof": "", "request_count": t, "total_requests": t, "approved_count": a, "rejected_count": rej, "unique_users": uniq, "offer_health": "", "approval_rate": f"{rate}%", "unique_requesters": uniq, "offer_total_requests": t, "days_since_request": days, "last_mail_sent": None, "publisher_account_age": "", "publisher_clicks": offer_clicks.get(oid, 0), "publisher_conversions": offer_convs.get(oid, 0), "message": ""})
        return results

    def _group_data(self, data, group_by):
        fm = {"user": "publisher_username", "network": "offer_network", "vertical": "offer_category", "country": "offer_countries"}
        field = fm.get(group_by, "offer_network")
        groups = {}
        for row in data:
            key = row.get(field, "") or "Unknown"
            groups.setdefault(key, []).append(row)
        return dict(sorted(groups.items(), key=lambda x: len(x[1]), reverse=True))

    def _format_value(self, col_key, value):
        if value is None: return ""
        if col_key in ("requested_at", "approved_at", "rejected_at", "last_mail_sent"):
            if isinstance(value, datetime): return value.strftime("%Y-%m-%d %H:%M")
            if isinstance(value, str) and value:
                d = self._parse_date(value)
                return d.strftime("%Y-%m-%d %H:%M") if d else value
        if col_key == "offer_payout" and isinstance(value, (int, float)): return f"${value:.2f}"
        return value

    def _write_sheet(self, ws, data, columns, freeze, color_code, auto_fit):
        hf = Font(bold=True, color=COLORS["header_font"], size=11)
        hfill = PatternFill(start_color=COLORS["header_bg"], end_color=COLORS["header_bg"], fill_type="solid")
        valid = [c for c in columns if c in ALL_COLUMNS]
        for ci, ck in enumerate(valid, 1):
            cell = ws.cell(row=1, column=ci, value=ALL_COLUMNS[ck]["label"])
            cell.font, cell.fill, cell.alignment = hf, hfill, Alignment(horizontal="center", vertical="center")
        for ri, rd in enumerate(data, 2):
            st = str(rd.get("status", "")).lower()
            for ci, ck in enumerate(valid, 1):
                cell = ws.cell(row=ri, column=ci, value=self._format_value(ck, rd.get(ck, "")))
                cell.alignment = Alignment(vertical="center")
                if color_code and st in STATUS_FILLS: cell.fill = STATUS_FILLS[st]
        if freeze: ws.freeze_panes = "A2"
        if auto_fit:
            for ci, ck in enumerate(valid, 1):
                ws.column_dimensions[get_column_letter(ci)].width = ALL_COLUMNS[ck]["width"]

    def _write_grouped_sheet(self, ws, groups, columns, group_by, freeze, color_code, auto_fit):
        hf = Font(bold=True, color=COLORS["header_font"], size=11)
        hfill = PatternFill(start_color=COLORS["header_bg"], end_color=COLORS["header_bg"], fill_type="solid")
        sf = Font(bold=True, color=COLORS["section_font"], size=11)
        sfill = PatternFill(start_color=COLORS["section_bg"], end_color=COLORS["section_bg"], fill_type="solid")
        valid = [c for c in columns if c in ALL_COLUMNS]
        for ci, ck in enumerate(valid, 1):
            cell = ws.cell(row=1, column=ci, value=ALL_COLUMNS[ck]["label"])
            cell.font, cell.fill, cell.alignment = hf, hfill, Alignment(horizontal="center", vertical="center")
        cr = 2
        for gn, gr in groups.items():
            cell = ws.cell(row=cr, column=1, value=f"{group_by.title()}: {gn} ({len(gr)} records)")
            cell.font, cell.fill = sf, sfill
            if len(valid) > 1: ws.merge_cells(start_row=cr, start_column=1, end_row=cr, end_column=len(valid))
            for c in range(1, len(valid)+1): ws.cell(row=cr, column=c).fill = sfill
            cr += 1
            for rd in gr:
                st = str(rd.get("status", "")).lower()
                for ci, ck in enumerate(valid, 1):
                    cell = ws.cell(row=cr, column=ci, value=self._format_value(ck, rd.get(ck, "")))
                    cell.alignment = Alignment(vertical="center")
                    if color_code and st in STATUS_FILLS: cell.fill = STATUS_FILLS[st]
                cr += 1
            cr += 1
        if freeze: ws.freeze_panes = "A2"
        if auto_fit:
            for ci, ck in enumerate(valid, 1):
                ws.column_dimensions[get_column_letter(ci)].width = ALL_COLUMNS[ck]["width"]

    def _create_summary_sheet(self, wb, summary_data, tabs, columns, config):
        ws = wb.create_sheet(title="Summary", index=0)
        ws.cell(row=1, column=1, value="Export Summary").font = Font(bold=True, size=16, color=COLORS["summary_title"])
        ws.cell(row=2, column=1, value=f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}")
        ws.cell(row=3, column=1, value=f"Tabs: {', '.join(tabs)}")
        ws.cell(row=4, column=1, value=f"Grouped by: {config.get('group_by', 'None')}")
        row = 6
        ws.cell(row=row, column=1, value="Tab").font = Font(bold=True)
        ws.cell(row=row, column=2, value="Records").font = Font(bold=True)
        ws.cell(row=row, column=3, value="Avg Payout").font = Font(bold=True)
        row += 1
        total = 0
        for tab in tabs:
            info = summary_data.get(tab, {})
            cnt = info.get("count", 0)
            total += cnt
            payouts = [r.get("offer_payout", 0) for r in info.get("data", []) if isinstance(r.get("offer_payout"), (int, float))]
            avg = sum(payouts)/len(payouts) if payouts else 0
            ws.cell(row=row, column=1, value=tab.replace("_", " ").title())
            ws.cell(row=row, column=2, value=cnt)
            ws.cell(row=row, column=3, value=f"${avg:.2f}")
            row += 1
        ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=row, column=2, value=total).font = Font(bold=True)
        row += 2
        # Top networks
        nets = {}
        for info in summary_data.values():
            for r in info.get("data", []):
                n = r.get("offer_network", "") or "Unknown"
                nets[n] = nets.get(n, 0) + 1
        if nets:
            ws.cell(row=row, column=1, value="Top Networks").font = Font(bold=True, size=12)
            row += 1
            for n, c in sorted(nets.items(), key=lambda x: x[1], reverse=True)[:10]:
                ws.cell(row=row, column=1, value=n)
                ws.cell(row=row, column=2, value=c)
                row += 1
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15

    def _safe_sheet_name(self, name):
        for ch in ["\\", "/", "*", "?", ":", "[", "]"]: name = name.replace(ch, "")
        return name[:31] or "Sheet"
